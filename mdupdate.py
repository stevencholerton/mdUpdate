# eFacs Method Desctiptions Updater
# Steve Cholerton
# Utopia Furniture Limited
# 3rd December October 2025

# Before Running - Install Libraries
#   $pip install openpyxl datetime SQLAlchemy pyodbc
# Make Sure Paths are Setup Appropriately

# HISTORY 
# 03/11/2025: Cloned from mdUpdate.py


import os
import openpyxl
import shutil
import datetime
import pyodbc
import sqlalchemy
import json
from sqlalchemy import create_engine
from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError

VERSION = "mdUpdate Version: 0.2 - 03/12/2025"
DEBUGMODE = False

# Clear the Terminal and Intro Credits
os.system('cls' if os.name == 'nt' else 'clear')
print("mdUpdate by Steven Cholerton")
print(VERSION)
print("SQLAlchemy Version: ", sqlalchemy.__version__, "\n")

def read_excel_columns(file_path):
    timestart = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    print(f"Starting Spreadsheet Validation: {timestart}")
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    if "Results" not in workbook.sheetnames:
        print(f"Results not found in the Spreadsheet: {file_path} ...")
        print("Please Ensure the Data to be Updated Is In the Sheet: Results ...\n")
        return
    sheet = workbook["Results"]

    row_number = 4 # First Row
    # Loop Rows, Starting at 4 (Excludes the Header Rows)
    # No Validation Required
    print("\nValidation Complete ...\n")
    return(row_number)


def update_efacs(file_path):
    print("Starting Update to eFacs ...\n")
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    if "Results" not in workbook.sheetnames:
        print(f"Results not found in the Spreadsheet: {file_path} ...")
        print("Please Ensure the Data to be Updated Is In the Sheet: Results ...\n")
        return
    sheet = workbook["Results"]

    row_number = 4 # First Row
    # Loop Rows, Starting at 3 (Excludes the Header Rows)
    total_updates = 0
    for row in sheet.iter_rows(min_row=4, min_col=1, max_col=11, values_only=True):
        if DEBUGMODE == True: print(f"UPDATING ROW: {row_number}")
        if row_number % 2000 == 0: print(f"ROWS UPDATED: {row_number}")
        partid = ""
        parameters = "{ "
        sql = "UPDATE [efacdb].[dbo].[stockedparts] SET "
        for col_index, value in enumerate(row, start=4):  # Columns D–K
            if col_index == 4:
                partid = f"{value}"
            if col_index == 7:
                sql = sql + f"minimumquantity = :minimumquantity, "
            if col_index == 8:
                sql = sql + f"maximumquantity = :maximumquantity, "
            if col_index == 9:
                sql = sql + f"reorderpoint = :reorderpoint, "
            if col_index == 10:
                sql = sql + f"reorderquantity = :reorderquantity, "
            if col_index == 11:
                sql = sql + f"roundingquantity = :roundingquantity, "
            if col_index == 12:
                sql = sql + f"leadtime = :leadtime, "
            if col_index == 13:
                sql = sql + f"reorderpolicy = :reorderpolicy, "
            if col_index == 14:
                sql = sql + f"usagerate = :usagerate "

            if col_index == 7:
                parameters = parameters + f'"minimumquantity": {value}, '
            if col_index == 8:
                parameters = parameters + f'"maximumquantity": {value}, '
            if col_index == 9:
                parameters = parameters + f'"reorderpoint": {value}, '
            if col_index == 10:
                parameters = parameters + f'"reorderquantity": {value}, '
            if col_index == 11:
                parameters = parameters + f'"roundingquantity": {value}, '
            if col_index == 12:
                parameters = parameters + f'"leadtime": {value}, '
            if col_index == 13:
                parameters = parameters + f'"reorderpolicy": "{value}", '
            if col_index == 14:
                parameters = parameters + f'"usagerate": "{value}", '

        sql = sql + f'WHERE partid = :partid'
        parameters = parameters + f'"partid": "{partid}"' + " }"
        engine = connect_db()
        if DEBUGMODE == True: print("SQL: " + sql)
        if DEBUGMODE == True: print("PARAMETERS: " + parameters)
        params_dict = json.loads(parameters)  # Convert String to Dict
        rows = perform_update(engine, sql, params_dict)
        if rows != 0:
            total_updates += 1
        if DEBUGMODE == True: print("-" * 40)  # Visual Separator
        row_number += 1
    return total_updates


def backup_file(file_path):
    if not os.path.isfile(file_path):
        print(f"Source File Not Found: {file_path}")
        return
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    base_name, ext = os.path.splitext(os.path.basename(file_path))  # Get Just Filename
    server_path = r"\\srv-prd-app\Filescan\LiveDB\Out\mdUpdate"
    new_file = os.path.join(server_path, f"{base_name}_{timestamp}{ext}")
    os.makedirs(server_path, exist_ok=True) # Ensure the Destination Folder Exists
    shutil.copy2(file_path, new_file)
    print(f"Backed Up: {file_path} to: {new_file}\n")


def delete_original_file(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"Original File Successfully Deleted: {file_path}\n")
    else:
        print(f"Original File Could Not be Deleted: {file_path}")
    

def connect_db():
    # Setup the Connection String
    connect_1 = "mssql+pyodbc://sn:=]VnGan2(m~LmM|MS[@192.168.1.180:1433/efacdb"
    connect_2 = "?driver=ODBC+Driver+17+for+SQL+Server"
    connection_string = connect_1 + connect_2
    # Setup the SQLAlchemy Engine
    engine = create_engine(connection_string)
    return (engine)


def perform_update(engine, sql: str, params: dict | None = None) -> int:
    if not isinstance(params, dict):
        raise TypeError("Params Must be a Dict of Bound Values ...")
    try:
        with engine.begin() as conn:
            result = conn.execute(text(sql), params or {})
            if DEBUGMODE == True: print(f"Update Executed Successfully, Rows Affected: {result.rowcount}")
            return result.rowcount or 0
    except SQLAlchemyError as e:
        print("Database Update Failed ...\nERROR: " ,e)
        return 0


# Program 
if __name__ == "__main__":
    file_path = "mdUpdate.xlsx"
    if os.path.exists(file_path):
        total_rows = read_excel_columns(file_path)
        total_updates = update_efacs(file_path)
        print(f"\n{total_updates} / {total_rows-4} Total Rows Updated ...")
        timeend = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        print(f"Time Finished: {timeend}")
        if os.name == 'nt': 
            backup_file(file_path)
            delete_original_file(file_path)
    else:
        print(f"File Could Not be Found: {file_path}\n")

